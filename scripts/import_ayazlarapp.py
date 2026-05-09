#!/usr/bin/env python3
import json
import sys
import urllib.request
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

API_URL = "https://script.google.com/macros/s/AKfycbyr8CyvG1ubgI9xUCR9esVIf7GLoSBFBtfXZCN3fNXB94QsvThEx6MX0qjVeO4nFZB3jw/exec"
SOURCE_PATH = Path("/Users/elifesrabora/Downloads/ayazlarapp.xlsx")


def iso(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return "" if value is None else str(value)


def rows(ws):
    headers = [cell.value for cell in ws[1]]
    clean_headers = [str(header).strip() if header else "" for header in headers]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and value != "" for value in row):
            continue
        yield {clean_headers[index]: value for index, value in enumerate(row) if clean_headers[index]}


def post(table, record):
    payload = {
        "action": "upsert",
        "table": table,
        "record": record,
        "user": {"email": "import@ayazlarapp.local", "role": "admin"},
    }
    request = urllib.request.Request(API_URL, data=json.dumps(payload).encode("utf-8"), method="POST")
    with urllib.request.urlopen(request, timeout=45) as response:
        result = json.loads(response.read().decode("utf-8"))
    if not result.get("ok"):
        raise RuntimeError(f"{table}: {result.get('error')}")
    return result


def main():
    wb = load_workbook(SOURCE_PATH, data_only=True)
    now = datetime.now().isoformat(timespec="seconds")
    imported = {key: 0 for key in ["projects", "sites", "tasks", "tasks_skipped", "reports", "personnel", "materials", "documents", "users"]}

    projects = list(rows(wb["Projects"]))
    project_locations = {item["id"]: item.get("location") or "" for item in projects}
    project_names = {item["id"]: item.get("name") or "" for item in projects}

    for item in projects:
        budget = item.get("budget") or 0
        location = item.get("location") or ""
        project_record = {
            "id": item["id"],
            "createdAt": now,
            "createdBy": "ayazlarapp-import",
            "name": item.get("name") or "",
            "client": "",
            "location": location,
            "startDate": iso(item.get("startDate")),
            "endDate": iso(item.get("endDate")),
            "budget": budget,
            "status": "Devam ediyor",
            "notes": f"Excel aktarımı. Konum: {location}. Bütçe: {budget}",
        }
        post("projects", project_record)
        imported["projects"] += 1

        post(
            "sites",
            {
                "id": f"site-{item['id']}",
                "createdAt": now,
                "createdBy": "ayazlarapp-import",
                "projectId": item["id"],
                "name": item.get("name") or "",
                "location": location,
                "manager": "",
                "status": "Aktif",
            },
        )
        imported["sites"] += 1

    for item in rows(wb["Reports"]):
        project_id = item.get("projectId") or ""
        report_date = iso(item.get("date"))
        attachment = item.get("FİKRETBEY APARTMANI GÜNLÜK SAHA RAPORU.pdf") or ""
        post(
            "reports",
            {
                "id": item.get("id") or f"report-{project_id}-{report_date}",
                "createdAt": iso(item.get("createdAt")) or now,
                "createdBy": item.get("createdById") or "ayazlarapp-import",
                "projectId": project_id,
                "siteId": f"site-{project_id}" if project_id else "",
                "date": report_date,
                "workingHours": item.get("workingHours") or "",
                "workDone": item.get("workSummary") or "",
                "nextPlan": item.get("nextPlan") or "",
                "incident": item.get("incident") or "",
                "notes": item.get("notes") or "",
                "attachmentName": attachment,
                "attachmentUrl": "",
            },
        )
        imported["reports"] += 1

    for item in rows(wb["Tasks"]):
        try:
            assigned_to = item.get("assignedToId") or ""
            post(
                "tasks",
                {
                    "id": item.get("id") or "",
                    "createdAt": iso(item.get("createdAt")) or now,
                    "createdBy": item.get("createdById") or "ayazlarapp-import",
                    "projectId": item.get("projectId") or "",
                    "title": item.get("title") or "",
                    "assignedTo": assigned_to,
                    "dueDate": iso(item.get("dueDate")),
                    "status": item.get("status") or "Planlandı",
                    "notes": item.get("note") or "",
                },
            )
            imported["tasks"] += 1
        except RuntimeError:
            imported["tasks_skipped"] += 1

    puantaj_dates = {item.get("id"): iso(item.get("date")) for item in rows(wb["Puantaj"])}
    for index, item in enumerate(rows(wb["Workers"]), start=1):
        puantaj_id = item.get("puantajId") or ""
        project_id = item.get("projectId") or ""
        status = "Geldi" if item.get("status") == "present" else "Gelmedi"
        post(
            "personnel",
            {
                "id": f"{puantaj_id}-{index}",
                "createdAt": now,
                "createdBy": "ayazlarapp-import",
                "projectId": project_id,
                "siteId": f"site-{project_id}" if project_id else "",
                "date": puantaj_dates.get(puantaj_id, ""),
                "name": item.get("name") or "",
                "job": item.get("job") or "",
                "attendance": status,
            },
        )
        imported["personnel"] += 1

    for item in rows(wb["Orders"]):
        quantity = item.get("quantity") or 0
        unit_price = item.get("unitPrice") or 0
        total = item.get("total") or (quantity * unit_price if quantity and unit_price else 0)
        post(
            "materials",
            {
                "id": item.get("id") or "",
                "createdAt": iso(item.get("createdAt")) or now,
                "createdBy": item.get("orderedById") or "ayazlarapp-import",
                "projectId": item.get("projectId") or "",
                "date": iso(item.get("date")),
                "name": item.get("material") or "",
                "spec": item.get("spec") or "",
                "quantity": quantity,
                "unit": item.get("unit") or "",
                "supplier": item.get("supplier") or "",
                "unitPrice": unit_price,
                "total": total,
                "minimum": "",
                "status": item.get("status") or "Sipariş verildi",
                "notes": item.get("note") or "",
            },
        )
        imported["materials"] += 1

    for item in rows(wb["Documents"]):
        project_id = item.get("projectId") or ""
        post(
            "documents",
            {
                "id": item.get("id") or "",
                "createdAt": iso(item.get("createdAt")) or now,
                "createdBy": item.get("createdById") or "ayazlarapp-import",
                "projectId": project_id,
                "siteId": f"site-{project_id}" if project_id else "",
                "title": item.get("title") or "",
                "type": item.get("type") or "",
                "fileName": item.get("title") or "",
                "fileUrl": item.get("url") or "",
                "mimeType": "",
                "notes": item.get("note") or "",
            },
        )
        imported["documents"] += 1

    for item in rows(wb["Users"]):
        username = item.get("username") or ""
        post(
            "users",
            {
                "id": item.get("id") or username,
                "createdAt": now,
                "createdBy": "ayazlarapp-import",
                "email": f"{username}@ayazlarapp.local" if username else "",
                "name": item.get("name") or "",
                "username": username,
                "role": item.get("role") or "viewer",
                "status": "active" if item.get("active") else "passive",
                "permissions": "*",
            },
        )
        imported["users"] += 1

    print(json.dumps(imported, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"Import failed: {exc}", file=sys.stderr)
        sys.exit(1)
